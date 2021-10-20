from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time


def waitUntilLoadIsInvis(seconds):
    WebDriverWait(browser, seconds).until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.loadingAnimation__text")))
    browser.execute_script('window.stop();')


def collect_referee_names(referees_refs_selectors):
    for selector in referees_refs_selectors:
        referee_refs.append(selector.get_attribute('href'))
        referee_names.append(selector.text)


t1 = time.time()
website_address = "https://www.transfermarkt.com/championship/schiedsrichter/wettbewerb/GB2"
LEAGUES = ('"Premier League"', '"LaLiga"', '"Serie A"', '"Ligue 1"', '"Bundesliga"', '"Premier Liga"', '"Championship"')
LEAGUE = LEAGUES[6]
REFEREE_REFS_SELECTOR = '//div[@id="yw1"]//td[@class="hauptlink"]/a'
SEASONS = ('2019', '2018', '2017', '2016', '2015', '2014', '2013', '2012', '2011', '2010')
FIRST_PART_SELECTOR = '//div[@class="table-header"]/img[@title='
SECOND_PART_SELECTOR = ']/../../div[@class="responsive-table"]//tbody/tr/td[@class="zentriert"]'
yellow_card_selector = f'{FIRST_PART_SELECTOR}{LEAGUE}{SECOND_PART_SELECTOR}[4]'
second_yellow_card_selector = f'{FIRST_PART_SELECTOR}{LEAGUE}{SECOND_PART_SELECTOR}[5]'
red_card_selector = f'{FIRST_PART_SELECTOR}{LEAGUE}{SECOND_PART_SELECTOR}[6]'
penalty_selector = f'{FIRST_PART_SELECTOR}{LEAGUE}{SECOND_PART_SELECTOR}[7]'
referee_refs, referee_names, full_links = [], [], []
temp_yellows, temp_second_yellows, temp_reds, temp_penalties, temp_stats = [], [], [], [], []
data = {}
options = webdriver.ChromeOptions()
options.add_argument('headless')
# caps = DesiredCapabilities.CHROME
browser = webdriver.Chrome(chrome_options=options) #, desired_capabilities=caps)
# caps["pageLoadStrategy"] = "none"
browser.implicitly_wait(0.5)
browser.get(website_address)
time.sleep(2)
referees_refs_selectors = browser.find_elements(By.XPATH, REFEREE_REFS_SELECTOR)
collect_referee_names(referees_refs_selectors)
browser.get(f'{website_address}/page/2')
time.sleep(2)
referees_refs_selectors = browser.find_elements(By.XPATH, REFEREE_REFS_SELECTOR)
collect_referee_names(referees_refs_selectors)
for i in range(len(referee_refs)):
    data.update({referee_names[i]: {}})
    print(referee_names[i])
    for year in SEASONS:
        print(year, end=' ')
        full_link = f'{referee_refs[i]}/saison/{year}'
        full_links.append(full_link)
        browser.get(full_link)
        time.sleep(1)
        temp_yellows_selectors = browser.find_elements(By.XPATH, yellow_card_selector)
        for selector in temp_yellows_selectors:
            temp_yellows.append(selector.text)
        temp_second_yellows_selectors = browser.find_elements(By.XPATH, second_yellow_card_selector)
        for selector in temp_second_yellows_selectors:
            temp_second_yellows.append(selector.text)
        temp_reds_selectors = browser.find_elements(By.XPATH, red_card_selector)
        for selector in temp_reds_selectors:
            temp_reds.append(selector.text)
        temp_penalties_selectors = browser.find_elements(By.XPATH, penalty_selector)
        for selector in temp_penalties_selectors:
            temp_penalties.append(selector.text)
        for j in zip(temp_yellows, temp_second_yellows, temp_reds, temp_penalties):
            temp_stats.append(j)
        data[referee_names[i]][f'{year[2:]}/{int(year[2:])+1}'] = temp_stats.copy()
        temp_stats.clear()
        temp_yellows.clear()
        temp_second_yellows.clear()
        temp_reds.clear()
        temp_penalties.clear()


browser.quit()

# print(data)
CELLS = ('Год', 'ЖК', '2-я ЖК', 'КК', 'Пен')
workbook = openpyxl.Workbook()
Sheet = workbook.get_sheet_by_name('Sheet')
workbook.remove_sheet(Sheet)
sheet_number = 0
for name in data.keys():
    workbook.create_sheet(str(name), sheet_number)
    workbook.active = sheet_number
    sheet = workbook.active
    column = 1
    for i in CELLS:
        sheet.cell(row=1, column=column, value=i)
        column += 1
    row = 2
    # print(diction[name])
    for key, value in data[name].items():
        sheet.cell(row=row, column=1, value=key)
        column = 2
        # print(value)
        for element in value:
            # print(element)
            for el in element:
                if el == '-':
                    sheet.cell(row=row, column=column, value='')
                else:
                    sheet.cell(row=row, column=column, value=int(el))
                column += 1
            row += 1
            column = 2
    sheet_number += 1
workbook.save(filename=f"{LEAGUE[1:-1]}.xlsx")

t2 = time.time()
print((t2-t1)/60)

